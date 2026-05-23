"""xlsx parser — openpyxl-driven, one `raw_pages` row per sheet.

Phase 2b. Per build_tracker §5.6 decisions:
- #1: one page per sheet (page_number = sheet index 1-based).
- #3: text format = `# Sheet: <name>\\n<row>\\n<row>...` where rows are
      tab-separated cell values.
- #6: magic-byte detection via ZIP signature `PK\\x03\\x04`.
- #13: empty sheets still emit a row (text may be just the header).
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from kb.parsers import Page, ParsedDocument, ParseError


_XLSX_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


class XLSXParser:
    """Spreadsheet → one page per sheet, tab-separated cells per row."""

    def can_handle(self, mime_type: str, magic_bytes: bytes) -> bool:
        if mime_type in _XLSX_MIMES:
            return True
        # ZIP magic — xlsx is a ZIP archive under the hood. The dispatcher
        # passes magic_bytes when mime is missing or generic; without an
        # explicit mime we treat ZIP-magic as xlsx (we don't accept raw .zip).
        return magic_bytes[:4] == b"PK\x03\x04"

    async def parse(
        self, file_bytes: bytes, *, file_id: str, workspace_id: str
    ) -> ParsedDocument:
        try:
            wb = load_workbook(
                filename=io.BytesIO(file_bytes),
                read_only=True,
                data_only=True,  # render formulas as their cached values
            )
        except Exception as exc:
            raise ParseError(f"openpyxl failed on file={file_id}: {exc}") from exc

        pages: list[Page] = []
        for idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            text, rows, cols = self._render_sheet(ws, sheet_name)
            pages.append(Page(
                page_number=idx,
                text=text,
                layout_json={
                    "sheet_name": sheet_name,
                    "rows": rows,
                    "cols": cols,
                },
            ))
        wb.close()

        if not pages:
            # An xlsx with zero sheets shouldn't happen (Excel requires ≥1),
            # but be defensive.
            raise ParseError(f"xlsx has no sheets: file={file_id}")
        return ParsedDocument(pages=pages)

    def _render_sheet(self, ws, sheet_name: str) -> tuple[str, int, int]:
        """Return (text, rows_count, cols_count). Text format:

            # Sheet: <name>
            cell\tcell\tcell
            cell\tcell\tcell

        Empty sheets return ('', 0, 0) — decision #13 (page row still emitted).
        """
        lines: list[str] = []
        row_count = 0
        col_count = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            row_values = ["" if v is None else str(v) for v in row]
            col_count = max(col_count, len(row_values))
            # Strip trailing all-empty padding so the TSV stays tight
            while row_values and row_values[-1] == "":
                row_values.pop()
            if not row_values:
                continue  # skip fully-empty rows from the text body
            lines.append("\t".join(row_values))

        if not lines:
            # Truly empty sheet — no header, no rows. Decision #13.
            return "", 0, 0

        body = "\n".join(lines)
        text = f"# Sheet: {sheet_name}\n{body}"
        return text, row_count, col_count
